"""
江戸川区専用 Excel編集サービス

江戸川区のExcelテンプレートに特化した処理を実装
"""
import openpyxl
from openpyxl.cell.cell import MergedCell
from pathlib import Path
from datetime import datetime, date
from typing import List, Dict, Optional

from services.excel_service_base import BaseExcelService


# 種別のカスタムソート順序
TYPE_ORDER = {
    "一般": 0,
    "35": 1,
    "45": 2,
    "55": 3,
    "65": 4,
}

# シングルス大会の種別名（オープン大会。会員登録表は不要で申込書テンプレートも専用）
SINGLES_TYPE = "シングルス"


class EdogawaExcelService(BaseExcelService):
    """江戸川区専用のExcel生成サービス"""

    def __init__(self):
        super().__init__(ward_id=23, ward_name="江戸川区")

    def _is_singles(self, tournament: Optional[Dict] = None,
                    registrations: Optional[List[Dict]] = None) -> bool:
        """シングルス大会かどうかを判定

        大会情報の type（JSONリスト）に "シングルス" が含まれる、
        または申込データの種別が "シングルス" であれば真。
        """
        if tournament is not None:
            t = tournament.get('type')
            types = t if isinstance(t, list) else ([t] if t else [])
            if SINGLES_TYPE in types:
                return True
        if registrations:
            return any(r.get('type') == SINGLES_TYPE for r in registrations)
        return False

    def _set_cell_value(self, ws, cell_ref, value):
        """
        セルに値を設定（結合セル対応）

        Args:
            ws: ワークシート
            cell_ref: セル参照（例: 'A1'）
            value: 設定する値
        """
        cell = ws[cell_ref]
        if not isinstance(cell, MergedCell):
            cell.value = value

    def generate_member_registration(
        self,
        tournament_name: str,
        registrations: List[Dict],
        timestamp: str
    ) -> Optional[Path]:
        """
        江戸川区の会員登録表を生成

        Args:
            tournament_name: 大会名
            registrations: 申込データリスト
            timestamp: タイムスタンプ

        Returns:
            生成されたファイルパス（会員登録対象者がいない・シングルス大会の場合はNone）
        """
        # シングルス大会はオープン大会のため会員登録表は出力しない
        if self._is_singles(registrations=registrations):
            return None

        # edogawa_flg=False の選手のみ抽出
        players_to_write = self._extract_unregistered_players(registrations)

        # 会員登録対象者がいなければ登録表は出力しない
        if not players_to_write:
            return None

        # テンプレートをコピー
        # ファイル名に大会名を含める（大会名ベースのクリーンアップ・一覧表示の対象にするため）
        output_filename = f"{tournament_name}_会員登録表.xlsx"
        output_path = self._copy_template("会員登録表フォーマット.xlsx", output_filename)

        # Excelを開く
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Sheet1"]

        # A2: 年度を設定（結合セルA2~H2）- 元のテキストの⚪︎を数字に置き換え
        current_value = ws['A2'].value or ""
        fiscal_year_number = self._get_fiscal_year_number()
        new_value = current_value.replace("⚪︎", str(fiscal_year_number))
        self._set_cell_value(ws, 'A2', new_value)

        # データ書き込み（6行目から）
        for idx, player in enumerate(players_to_write):
            row = 6 + idx

            # A列: No
            self._set_cell_value(ws, f'A{row}', idx + 1)

            # BC列: 氏名（結合セル）
            self._set_cell_value(ws, f'B{row}', player['player_name'])

            # D列: 生年月日（YYYY/MM/DD形式の文字列で出力）
            birth_date = player['birth_date']
            if isinstance(birth_date, str):
                birth_date = datetime.strptime(birth_date, "%Y-%m-%d").date()
            self._set_cell_value(ws, f'D{row}', birth_date.strftime("%Y/%m/%d") if birth_date else "")

            # E列: 年齢（計算式を維持するため、ここでは触らない）
            # ※テンプレートに既に計算式が入っているため

            # F列: 性別
            sex = player['sex']
            self._set_cell_value(ws, f'F{row}', "男" if sex == 0 else "女")

            # G列: 郵便番号
            self._set_cell_value(ws, f'G{row}', player.get('post_number', ''))

            # H列: 現住所（結合セルH-J）
            self._set_cell_value(ws, f'H{row}', player['address'])

            # KL列: 電話（結合セルK-L）
            self._set_cell_value(ws, f'K{row}', player['phone_number'])

        # 保存
        wb.save(output_path)
        wb.close()

        return output_path

    def generate_individual_application(
        self,
        tournament: Dict,
        registrations: List[Dict],
        timestamp: str
    ) -> Optional[Path]:
        """
        江戸川区の個人戦申込書を生成

        Args:
            tournament: 大会情報
            registrations: 申込データリスト
            timestamp: タイムスタンプ

        Returns:
            生成されたファイルパス
        """
        # シングルス大会は専用テンプレート・レイアウトで生成
        if self._is_singles(tournament=tournament, registrations=registrations):
            return self._generate_singles_application(tournament, registrations, timestamp)

        # テンプレートをコピー
        output_filename = f"{tournament['tournament_name']}_申込書.xlsx"
        output_path = self._copy_template("個人戦_申込書フォーマット.xlsx", output_filename)

        # Excelを開く
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Sheet1"]

        # A1: 大会名を設定
        self._set_cell_value(ws, 'A1', f"{tournament['tournament_name']}　申込書")

        # F3: 申込日を設定
        self._set_cell_value(ws, 'F3', self._format_application_date())

        # 団体名・責任者・連絡先（5-7行目）は空欄のまま（手動入力）

        # 申込データをカスタムソート
        sorted_registrations = self._sort_registrations(registrations)

        # データ書き込み（10行目から）
        row = 10
        entry_no = 1

        for reg in sorted_registrations:
            # 種別文字列の生成
            type_str = self._format_type_sex(reg['type'], reg['sex'])

            # 申込者とペア相手のデータ
            applicant = reg.get('applicant')
            partner = reg.get('partner')

            # 2行書き込み（申込者 + ペア）
            players = [applicant, partner]

            for idx, player in enumerate(players):
                if player is None:
                    continue

                # A列: NO（1行目のみ書き込み - 結合セル対応）
                if idx == 0:  # 申込者の行のみ
                    self._set_cell_value(ws, f'A{row}', entry_no)

                # B列: 種別
                self._set_cell_value(ws, f'B{row}', type_str)

                # C列: 氏名
                self._set_cell_value(ws, f'C{row}', player['player_name'])

                # D列: 生年月日（YYYY/MM/DD形式の文字列で出力）
                birth_date = player['birth_date']
                if isinstance(birth_date, str):
                    birth_date = datetime.strptime(birth_date, "%Y-%m-%d").date()
                self._set_cell_value(ws, f'D{row}', birth_date.strftime("%Y/%m/%d") if birth_date else "")

                # E列: 所属名
                self._set_cell_value(ws, f'E{row}', player.get('affiliated_club', ''))

                # F列: 備考は空欄

                row += 1

            entry_no += 1

        # 保存
        wb.save(output_path)
        wb.close()

        return output_path

    # テンプレートの申込欄（先頭行のみ書き込む2行結合ブロック）
    SINGLES_TEMPLATE = "シングルス_申込書フォーマット.xlsx"
    SINGLES_START_ROW = 10   # 最初の申込欄の先頭行
    SINGLES_ROW_STRIDE = 2   # 1申込あたり2行（A/B列が2行結合）
    SINGLES_MAX_ENTRIES = 13  # テンプレートの申込欄数（10,12,...,34行）

    def _generate_singles_application(
        self,
        tournament: Dict,
        registrations: List[Dict],
        timestamp: str
    ) -> Optional[Path]:
        """
        江戸川区シングルス大会の申込書を生成（専用テンプレート）

        1申込 = 1選手（ペアなし）。テンプレートは1申込あたり2行結合ブロックのため、
        各申込を各ブロックの先頭行に記入する。
        """
        # テンプレートをコピー
        output_filename = f"{tournament['tournament_name']}_申込書.xlsx"
        output_path = self._copy_template(self.SINGLES_TEMPLATE, output_filename)

        # Excelを開く
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Sheet1"]

        # A1: 大会名 / F3: 申込日
        self._set_cell_value(ws, 'A1', f"{tournament['tournament_name']}　申込書")
        self._set_cell_value(ws, 'F3', self._format_application_date())

        # 男子→女子の順（各性別内は申込順を維持＝安定ソート）
        ordered = sorted(registrations, key=lambda r: (r.get('sex') or 0))

        entry_no = 0
        for reg in ordered:
            applicant = reg.get('applicant')
            if not applicant:
                continue

            row = self.SINGLES_START_ROW + entry_no * self.SINGLES_ROW_STRIDE
            if entry_no >= self.SINGLES_MAX_ENTRIES:
                print(f"⚠️ シングルス申込がテンプレート上限({self.SINGLES_MAX_ENTRIES}件)を"
                      f"超過したため以降を省略: {len(ordered)}件")
                break
            entry_no += 1

            # A列: NO
            self._set_cell_value(ws, f'A{row}', entry_no)

            # B列: 種別（例: シングルス男子 / シングルス女子）
            self._set_cell_value(ws, f'B{row}', self._format_type_sex(reg.get('type', SINGLES_TYPE), reg.get('sex')))

            # C列: 氏名
            self._set_cell_value(ws, f'C{row}', applicant['player_name'])

            # D列: 生年月日（YYYY/MM/DD形式の文字列で出力）
            birth_date = applicant['birth_date']
            if isinstance(birth_date, str):
                birth_date = datetime.strptime(birth_date, "%Y-%m-%d").date()
            self._set_cell_value(ws, f'D{row}', birth_date.strftime("%Y/%m/%d") if birth_date else "")

            # E列: 所属名
            self._set_cell_value(ws, f'E{row}', applicant.get('affiliated_club', ''))

            # F列: 備考は空欄

        # 保存
        wb.save(output_path)
        wb.close()

        return output_path

    # --- 江戸川区固有のヘルパーメソッド ---

    def _extract_unregistered_players(self, registrations: List[Dict]) -> List[Dict]:
        """
        edogawa_flg=False の選手のみ抽出（江戸川区特有）

        Args:
            registrations: 申込データリスト

        Returns:
            未登録選手のリスト（重複除去済み）
        """
        players_to_write = []

        for reg in registrations:
            # 申込者本人（discord_id に紐づく選手）
            applicant = reg.get('applicant')
            if applicant and not applicant.get('edogawa_flg', True):
                players_to_write.append(applicant)

            # ペア相手（pair1 に紐づく選手）
            partner = reg.get('partner')
            if partner and not partner.get('edogawa_flg', True):
                players_to_write.append(partner)

        # 重複除去（同じ選手が複数回出てくる可能性）
        unique_players = {}
        for player in players_to_write:
            player_id = player['player_id']
            if player_id not in unique_players:
                unique_players[player_id] = player

        return list(unique_players.values())

    def _sort_registrations(self, registrations: List[Dict]) -> List[Dict]:
        """
        申込データをカスタムソート（江戸川区特有）

        順序: 一般男子 → 35男子 → 45男子 → 一般女子 → 35女子 → 45女子

        Args:
            registrations: 申込データリスト

        Returns:
            ソートされた申込データリスト
        """
        return sorted(
            registrations,
            key=lambda r: (
                r['sex'],  # 0(男子)が先、1(女子)が後
                TYPE_ORDER.get(r['type'], 999)  # 一般→35→45→55...
            )
        )

    def _format_type_sex(self, type_: str, sex: int) -> str:
        """
        種別と性別を結合して文字列化

        Args:
            type_: 種別（"一般", "35", "45"など）
            sex: 性別（0:男子, 1:女子）

        Returns:
            結合文字列（例: "一般男子", "35女子"）
        """
        sex_str = "男子" if sex == 0 else "女子"
        return f"{type_}{sex_str}"

    def _format_application_date(self) -> str:
        """
        申込日を江戸川区フォーマットで生成

        Returns:
            申込日文字列（例: "申　込　日　令和7年11月18日"）
        """
        today = datetime.now()
        reiwa_year = today.year - 2018
        return f"申　込　日　令和{reiwa_year}年{today.month:02d}月{today.day:02d}日"

    def _get_fiscal_year_number(self) -> int:
        """
        現在の年度を令和年数で取得

        Returns:
            令和年数（例: 7）
        """
        today = datetime.now()

        # 年度計算（4月以降は今年、1-3月は前年）
        if today.month >= 4:
            fiscal_year = today.year
        else:
            fiscal_year = today.year - 1

        reiwa_year = fiscal_year - 2018
        return reiwa_year
