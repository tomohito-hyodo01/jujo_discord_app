"""
荒川区 Excel生成サービス

申込書のみ生成（個人登録表は不要）。
1ペア=2行のフォーマット。
"""
import openpyxl
from pathlib import Path
from datetime import datetime, date
from typing import List, Dict, Optional

from services.excel_service_base import BaseExcelService


# 種別ソート順序
TYPE_ORDER = {
    "一般": 0,
    "35": 1,
    "45": 2,
    "55": 3,
    "65": 4,
}


class ArakawaExcelService(BaseExcelService):
    """荒川区専用のExcel生成サービス"""

    TEMPLATE_FILENAME = "申込書フォーマット.xlsx"
    MAX_PRESET_PAIRS = 11  # テンプレートにあらかじめ番号が入っている行数

    def __init__(self):
        super().__init__(ward_id=18, ward_name="荒川区")

    # generate_member_registration は基底クラスのデフォルト (None) をそのまま使用

    def generate_individual_application(
        self,
        tournament: Dict,
        registrations: List[Dict],
        timestamp: str
    ) -> Optional[Path]:
        """
        荒川区フォーマットの申込書 (Excel) を生成
        """
        tournament_name = tournament.get('tournament_name', '')
        safe_name = self._safe_filename(tournament_name)
        output_filename = f"{safe_name}_申込書_{timestamp}.xlsx"

        output_path = self._copy_template(self.TEMPLATE_FILENAME, output_filename)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # A1 タイトル: "■{大会名} 申込ペア"
        ws['A1'] = f"■{tournament_name} 申込ペア"

        sorted_regs = self._sort_registrations(registrations)
        ref_date = self._parse_date(tournament.get('tournament_date'))

        for i, reg in enumerate(sorted_regs, start=1):
            row_top = 3 + (i - 1) * 2  # 1:3, 2:5, 3:7 ...
            row_bot = row_top + 1

            applicant = reg.get('applicant') or {}
            partner = reg.get('partner') or {}

            applicant_sex = applicant.get('sex', 0) or 0
            type_str = self._format_type_sex(reg.get('type', ''), applicant_sex)

            # 種別ごとに開催日が異なる大会（統合された兄弟レコード）では、
            # その申込が属するレコードの開催日を年齢基準日にする
            row_ref_date = self._parse_date(reg.get('source_tournament_date')) or ref_date

            # 上段: ペア番号 / 種別 / 申込者氏名 / 申込者年齢
            ws.cell(row_top, 1).value = i
            ws.cell(row_top, 2).value = type_str
            ws.cell(row_top, 3).value = self._normalize_name(applicant.get('player_name', ''))
            ws.cell(row_top, 4).value = self._calc_age(applicant.get('birth_date'), row_ref_date)

            # 下段: ペア相手氏名 / ペア相手年齢
            if partner:
                ws.cell(row_bot, 3).value = self._normalize_name(partner.get('player_name', ''))
                ws.cell(row_bot, 4).value = self._calc_age(partner.get('birth_date'), row_ref_date)

        # テンプレートに残っている余分なペア番号(1〜11)をクリア
        used_pairs = len(sorted_regs)
        if used_pairs < self.MAX_PRESET_PAIRS:
            for j in range(used_pairs + 1, self.MAX_PRESET_PAIRS + 1):
                row = 3 + (j - 1) * 2
                ws.cell(row, 1).value = None

        wb.save(output_path)
        return output_path

    def _sort_registrations(self, registrations: List[Dict]) -> List[Dict]:
        """男子→女子の順、内部は 一般→35→45→55→65 の順"""
        return sorted(
            registrations,
            key=lambda r: (
                (r.get('applicant') or {}).get('sex', 0) or 0,
                TYPE_ORDER.get(r.get('type', ''), 999),
                (r.get('applicant') or {}).get('player_name', '') or '',
            )
        )

    def _format_type_sex(self, type_: str, sex: int) -> str:
        sex_str = "男子" if sex == 0 else "女子"
        return f"{type_}{sex_str}"

    def _normalize_name(self, name: str) -> str:
        """姓と名の間の半角スペースを全角に統一"""
        if not name:
            return ''
        return name.strip().replace(' ', '　')

    def _calc_age(self, birth_date, ref_date) -> Optional[int]:
        """ref_date時点での年齢を計算"""
        if not birth_date or not ref_date:
            return None
        bd = self._to_date(birth_date)
        rd = self._to_date(ref_date)
        if bd is None or rd is None:
            return None
        age = rd.year - bd.year
        if (rd.month, rd.day) < (bd.month, bd.day):
            age -= 1
        return age

    def _parse_date(self, value):
        return self._to_date(value)

    def _to_date(self, value):
        if value is None:
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S'):
                try:
                    return datetime.strptime(value[:len(fmt)], fmt).date()
                except ValueError:
                    continue
            try:
                return datetime.fromisoformat(value.replace(' ', 'T')).date()
            except ValueError:
                return None
        return None

    def _safe_filename(self, name: str) -> str:
        """ファイル名に使えない文字を除去"""
        bad = '/\\:*?"<>|'
        return ''.join(c for c in (name or '') if c not in bad).strip()
