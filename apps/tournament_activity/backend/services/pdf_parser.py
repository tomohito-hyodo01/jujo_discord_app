#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
大会要項解析サービス

Claude API または Gemini API を使用して大会要項(PDF/画像/Excel)から情報を抽出
"""

import os
import base64
import json
from typing import Dict, Any, Literal, Optional
import anthropic


# ファイル拡張子 → MIMEタイプ
MIME_MAP = {
    '.pdf': 'application/pdf',
    '.jpg': 'image/jpeg',
    '.jpeg': 'image/jpeg',
    '.png': 'image/png',
    '.gif': 'image/gif',
    '.webp': 'image/webp',
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    '.xls': 'application/vnd.ms-excel',
}

IMAGE_MIMES = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
EXCEL_MIMES = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/vnd.ms-excel',
}


def _detect_mime(filename: Optional[str], content: bytes) -> str:
    """ファイル名またはマジックバイトからMIMEタイプを推定"""
    if filename:
        ext = os.path.splitext(filename)[1].lower()
        if ext in MIME_MAP:
            return MIME_MAP[ext]
    # マジックバイトで判定
    if content[:4] == b'%PDF':
        return 'application/pdf'
    if content[:8] == b'\x89PNG\r\n\x1a\n':
        return 'image/png'
    if content[:2] == b'\xff\xd8':
        return 'image/jpeg'
    if content[:4] == b'PK\x03\x04':
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return 'application/pdf'


def _extract_excel_text(content: bytes) -> str:
    """Excelファイルからテキストを抽出"""
    import openpyxl
    import io
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"=== シート: {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else '' for c in row]
            if any(cells):
                lines.append('\t'.join(cells))
    return '\n'.join(lines)


class PDFParserService:
    """大会要項解析サービス（Claude API または Gemini API 使用）"""

    def __init__(self, model: Literal["claude", "gemini"] = "claude"):
        self.model = model

        if model == "claude":
            self.api_key = os.getenv("ANTHROPIC_API_KEY")
            if not self.api_key:
                raise ValueError("ANTHROPIC_API_KEY is not set")
            self.client = anthropic.Anthropic(api_key=self.api_key)
        elif model == "gemini":
            self.api_key = os.getenv("GEMINI_API_KEY")
            if not self.api_key:
                raise ValueError("GEMINI_API_KEY is not set")
            import google.generativeai as genai
            genai.configure(api_key=self.api_key)
            self.client = genai.GenerativeModel('gemini-2.5-flash')
        else:
            raise ValueError(f"Unsupported model: {model}")

    async def parse_tournament_pdf(self, pdf_content: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
        """
        ファイルから大会情報を抽出（PDF/画像/Excel対応）
        """
        mime = _detect_mime(filename, pdf_content)

        if mime in EXCEL_MIMES:
            return await self._parse_excel(pdf_content)
        elif self.model == "claude":
            return await self._parse_with_claude(pdf_content, mime)
        elif self.model == "gemini":
            return await self._parse_with_gemini(pdf_content, mime)
        else:
            raise ValueError(f"Unsupported model: {self.model}")

    async def _parse_excel(self, content: bytes) -> Dict[str, Any]:
        """Excelをテキスト抽出してAIで解析"""
        text = _extract_excel_text(content)
        prompt = self._get_extraction_prompt()

        if self.model == "claude":
            message = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=2048,
                messages=[{
                    "role": "user",
                    "content": f"以下は大会要項のExcelデータです。\n\n{text}\n\n{prompt}"
                }]
            )
            return self._extract_json_from_response(message.content[0].text.strip())
        else:
            response = self.client.generate_content(
                f"以下は大会要項のExcelデータです。\n\n{text}\n\n{prompt}"
            )
            return self._extract_json_from_response(response.text.strip())

    async def _parse_with_claude(self, content: bytes, mime: str) -> Dict[str, Any]:
        """Claude APIを使用して解析（PDF/画像）"""
        content_base64 = base64.standard_b64encode(content).decode('utf-8')
        prompt = self._get_extraction_prompt()

        if mime in IMAGE_MIMES:
            doc_block = {
                "type": "image",
                "source": {"type": "base64", "media_type": mime, "data": content_base64}
            }
        else:
            doc_block = {
                "type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": content_base64}
            }

        message = self.client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2048,
            messages=[{
                "role": "user",
                "content": [doc_block, {"type": "text", "text": prompt}]
            }]
        )
        return self._extract_json_from_response(message.content[0].text.strip())

    async def _parse_with_gemini(self, content: bytes, mime: str) -> Dict[str, Any]:
        """Gemini APIを使用して解析（PDF/画像）"""
        prompt = self._get_extraction_prompt()
        response = self.client.generate_content([
            {"mime_type": mime, "data": content},
            prompt
        ])
        return self._extract_json_from_response(response.text.strip())

    def _get_extraction_prompt(self) -> str:
        """PDF情報抽出用のプロンプトを取得"""
        return """このPDFは大会要項です。以下の情報を抽出してJSON形式で返してください。

**重要**: 種別によって開催日が異なる場合は、開催日ごとに別のレコードとして返してください。
その場合、JSON配列として複数のオブジェクトを返してください。

例: 一般が9/13、35・45が9/6、ミックスが9/20の場合:
[
  {"tournament_name": "大会名", "tournament_date": "2026-09-06", "type": ["35", "45"], ...},
  {"tournament_name": "大会名", "tournament_date": "2026-09-13", "type": ["一般"], ...},
  {"tournament_name": "大会名", "tournament_date": "2026-09-20", "type": ["ミックス（一般）"], ...}
]

全種別が同じ日程の場合は、単一オブジェクトで返してください（配列不要）。

各レコードのフィールド:
{
  "tournament_id": "大会名と日付からMD5ハッシュで自動生成（例: tournament_a1b2c3d4）",
  "tournament_name": "大会名（完全な名称）",
  "registrated_ward": 主催区のID（下記参照）,
  "deadline_date": "申込締切日（YYYY-MM-DD形式、例: 2024-03-15）",
  "tournament_date": "大会開催日（YYYY-MM-DD形式、例: 2024-03-20）",
  "classification": 競技形式（個人戦=0, 団体戦=1）,
  "mix_flg": false,
  "type": ["一般", "35", "ミックス（一般）", "ミックス（35）"],
  "venue": "会場名（記載がなければnull）",
  "reception_time": "受付時刻（HH:MM形式、例: 08:15。記載がなければnull）",
  "opening_time": "開会式時刻（HH:MM形式、例: 08:45。記載がなければnull）",
  "match_start_time": "試合開始時刻（HH:MM形式。記載がなければnull）",
  "entry_fee": "参加費（例: 1ペア 2,000円。記載がなければnull）"
}

主催区のIDマッピング:
- 江戸川区 = 23
- 荒川区 = 18
- 北区 = 17
- 江東区 = 8
- 中央区 = 2
- 墨田区 = 7
- 文京区 = 5
- 浦安市 = 100
- 流山市 = 101
- 広域（東京都、関東、東日本、全日本など） = 99
- 上記以外の区・該当なし = -1 (エラーとして扱われます)

重要な注意事項:
1. 日付は必ずYYYY-MM-DD形式に変換してください（和暦の場合は西暦に変換）
2. tournament_idは必ず null にしてください（システム側で一意なIDを自動生成します。AIは生成しないこと）
3. 主催が「東京都」「関東」「東日本」「全日本」などの広域の場合は registrated_ward = 99 を設定してください
4. 主催区が上記リストにない場合（例：世田谷区、練馬区など）は registrated_ward = -1 を設定してください
5. **重要**: 種別(type)には「一般」「35」「45」「ミックス（一般）」「ミックス（35）」「シングルス」のみを抽出してください。ミックスダブルスは年齢区分に応じて「ミックス（一般）」または「ミックス（35）」としてください。シングルスの部がある場合は「シングルス」としてください。「シニア」「55」「60」「65」「70」「シニアミックス」などは無視してください
6. mix_flgは常にfalseにしてください（ミックスはtype配列で管理します）
7. JSONのみを返し、説明文やマークダウンは不要です

JSON:"""

    def _validate_tournament(self, tournament_data: Dict[str, Any]) -> Dict[str, Any]:
        """個別の大会データを検証・補正"""
        import hashlib

        # 主催区の検証
        allowed_ward_ids = {2, 5, 7, 8, 17, 18, 23, 99, 100, 101}
        ward_id = tournament_data.get("registrated_ward", 0)

        if ward_id == -1 or (ward_id != 0 and ward_id not in allowed_ward_ids):
            raise ValueError("INVALID_WARD")

        # tournament_idは必ずサーバー側で生成する（AIが返した値は信頼しない）。
        # AIはMD5を計算できず、それらしい16進文字列を捏造するため、別大会同士でIDが
        # 衝突し、register_tournamentのUPDATEで既存大会が上書き・消失する事故が起きていた。
        # 主催区＋大会名＋日付からMD5を計算することで、区・大会ごとに必ず一意なIDとなる。
        ward_for_id = tournament_data.get("registrated_ward", 0)
        id_source = f"{ward_for_id}_{tournament_data.get('tournament_name', 'unknown')}_{tournament_data.get('tournament_date', '')}"
        hash_value = hashlib.md5(id_source.encode()).hexdigest()[:8]
        tournament_data["tournament_id"] = f"tournament_{hash_value}"

        # 種別のフィルタリング：一般、35、45、ミックスのみを許可
        if "type" in tournament_data and isinstance(tournament_data["type"], list):
            allowed_types = {"一般", "35", "45", "ミックス（一般）", "ミックス（35）", "シングルス"}
            filtered_types = [
                t for t in tournament_data["type"]
                if t in allowed_types or (t.isdigit() and int(t) < 55)
            ]
            tournament_data["type"] = filtered_types

        return tournament_data

    def _extract_json_from_response(self, response_text: str):
        """AIレスポンスからJSONを抽出してパース。単一または配列を返す"""
        # マークダウンコードブロックを削除（もしあれば）
        if response_text.startswith('```'):
            response_text = response_text.split('```')[1]
            if response_text.startswith('json'):
                response_text = response_text[4:]
            response_text = response_text.strip()

        # JSONをパース
        parsed = json.loads(response_text)

        # 配列の場合は各要素を検証
        if isinstance(parsed, list):
            return [self._validate_tournament(item) for item in parsed]

        # 単一オブジェクトの場合
        return self._validate_tournament(parsed)
