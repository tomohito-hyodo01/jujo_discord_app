#!/usr/bin/env python3
"""
Excelテンプレートファイルの詳細構造を解析するスクリプト
"""
import openpyxl
from pathlib import Path

def analyze_excel_detailed(file_path):
    """Excelファイルの詳細構造を解析"""
    print(f"\n{'='*80}")
    print(f"📄 ファイル: {Path(file_path).name}")
    print('='*80)

    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        print(f"\n📊 シート名: {sheet_name}")
        print(f"   サイズ: {ws.max_row}行 x {ws.max_column}列")
        print(f"\n   最初の20行の内容:")
        print("   " + "-" * 76)

        # 最初の20行を表示
        for row in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(15, ws.max_column + 1)):  # 最初の15列まで
                cell = ws.cell(row=row, column=col)
                value = str(cell.value) if cell.value is not None else ""
                # 長すぎる値は切り詰め
                if len(value) > 15:
                    value = value[:12] + "..."
                row_data.append(value)

            # 空行でない場合のみ表示
            if any(row_data):
                col_letter = openpyxl.utils.get_column_letter(1)
                print(f"   {row:2d}行目: {row_data}")

        # 結合セル情報
        if ws.merged_cells:
            print(f"\n   結合セル: {len(ws.merged_cells.ranges)}個")
            for merged_range in list(ws.merged_cells.ranges)[:5]:  # 最初の5個のみ表示
                print(f"      {merged_range}")

    wb.close()

def main():
    template_dir = Path("apps/tournament_activity/backend/templates/wards/23_江戸川区")

    excel_files = [
        "会員登録表フォーマット.xlsx",
        "個人戦_申込書フォーマット.xlsx",
        "団体戦_申込書フォーマット.xlsx"
    ]

    for filename in excel_files:
        file_path = template_dir / filename
        if file_path.exists():
            analyze_excel_detailed(file_path)
        else:
            print(f"❌ ファイルが見つかりません: {filename}")

if __name__ == "__main__":
    main()
