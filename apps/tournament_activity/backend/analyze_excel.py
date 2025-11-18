#!/usr/bin/env python3
"""
Excelテンプレートファイルの構造を解析するスクリプト
"""
import openpyxl
from pathlib import Path
import json

def analyze_excel(file_path):
    """Excelファイルの構造を解析"""
    wb = openpyxl.load_workbook(file_path)

    result = {
        "file_name": Path(file_path).name,
        "sheets": []
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # ヘッダー行を取得（1行目）
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append({
                    "column": openpyxl.utils.get_column_letter(col),
                    "value": str(cell_value)
                })

        # データ例（2-3行目）
        data_rows = []
        for row in range(2, min(4, ws.max_row + 1)):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else "")
            if any(row_data):  # 空行でない場合のみ追加
                data_rows.append(row_data)

        sheet_info = {
            "sheet_name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "headers": headers,
            "data_examples": data_rows
        }

        result["sheets"].append(sheet_info)

    wb.close()
    return result

def main():
    template_dir = Path("apps/tournament_activity/backend/templates/wards/23_江戸川区")

    excel_files = list(template_dir.glob("*.xlsx"))

    if not excel_files:
        print("❌ Excelファイルが見つかりません")
        return

    results = []

    for excel_file in sorted(excel_files):
        print(f"\n{'='*60}")
        print(f"📄 ファイル: {excel_file.name}")
        print('='*60)

        try:
            result = analyze_excel(excel_file)
            results.append(result)

            for sheet in result["sheets"]:
                print(f"\n📊 シート名: {sheet['sheet_name']}")
                print(f"   行数: {sheet['max_row']}, 列数: {sheet['max_column']}")
                print(f"\n   ヘッダー行:")
                for header in sheet["headers"]:
                    print(f"      {header['column']}列: {header['value']}")

                if sheet["data_examples"]:
                    print(f"\n   データ例:")
                    for idx, row_data in enumerate(sheet["data_examples"], start=2):
                        print(f"      {idx}行目: {row_data[:5]}...")  # 最初の5列のみ表示

        except Exception as e:
            print(f"❌ エラー: {e}")

    # JSON形式で保存
    output_file = template_dir / "structure_analysis.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"\n\n✅ 解析結果を保存しました: {output_file}")

if __name__ == "__main__":
    main()
